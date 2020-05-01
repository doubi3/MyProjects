# import openpyxl and tkinter modules 
from openpyxl import *
from tkinter import *

# globally declare wb and sheet variable 

# opening the existing excel file 
wb = load_workbook('C:\\Users\\Mark Ekperi\\MyProjects\\RegistrationForm\\StudentRegForm.xlsx') 

# create the sheet object 
sheet = wb.active 


def excel(): 
	
	# resize the width of columns in 
	# excel spreadsheet 
	sheet.column_dimensions['A'].width = 20
	sheet.column_dimensions['B'].width = 20
	sheet.column_dimensions['C'].width = 20
	sheet.column_dimensions['D'].width = 20
	sheet.column_dimensions['E'].width = 20
	sheet.column_dimensions['F'].width = 20
	sheet.column_dimensions['G'].width = 20
	sheet.column_dimensions['H'].width = 20
	sheet.column_dimensions['I'].width = 20
	sheet.column_dimensions['J'].width = 20
	sheet.column_dimensions['K'].width = 20
	sheet.column_dimensions['L'].width = 20
	sheet.column_dimensions['M'].width = 20






	# write given data to an excel spreadsheet 
	# at particular location 
	sheet.cell(row=1, column=1).value = "Surname"
	sheet.cell(row=1, column=2).value = "Other Names"
	sheet.cell(row=1, column=3).value = "Mat/Registration Number"
	sheet.cell(row=1, column=4).value = "Year Admitted"
	sheet.cell(row=1, column=5).value = "Level"
	sheet.cell(row=1, column=6).value = "Mobile Number"
	sheet.cell(row=1, column=7).value = "Sex"
	sheet.cell(row=1, column=8).value = "Address"
	sheet.cell(row=1, column=9).value = "State"
	sheet.cell(row=1, column=10).value = "LGA"
	sheet.cell(row=1, column=11).value = "Parents/Guardian"
	sheet.cell(row=1, column=12).value = "Parent's Phone"
	sheet.cell(row=1, column=13).value = "Student's Email"



# Function to set focus (cursor) 

def focus1(event):

	surname.focus_set()
	
def focus2(event): 
	# set focus on the Other_Names box 
	other_names.focus_set() 


# Function to set focus 
def focus3(event): 
	# set focus on the Mat_Reg_No box 
	mat_reg_no.focus_set() 


# Function to set focus 
def focus4(event): 
	# set focus on the Year_Admitted box 
	year_admitted.focus_set() 


# Function to set focus 
def focus5(event): 
	# set focus on the Level box 
	level.focus_set() 


# Function to set focus 
def focus6(event): 
	# set focus on the Mobile_Number box 
	mobile_number.focus_set() 


# Function to set focus 
def focus7(event): 
	# set focus on the Sex box 
	sex.focus_set() 

# Function to set focus (cursor) 
def focus8(event): 
	# set focus on the Address box 
	address.focus_set() 


# Function to set focus 
def focus9(event): 
	# set focus on the State box 
	state.focus_set() 


# Function to set focus 
def focus10(event): 
	# set focus on the LGA box 
	lga.focus_set() 


# Function to set focus 
def focus11(event): 
	# set focus on the Parents_Guardian box 
	parents_guardian.focus_set() 


# Function to set focus 
def focus12(event): 
	# set focus on the Parent_Phone box 
	parent_phone.focus_set() 


# Function to set focus 
def focus13(event): 
	# set focus on the Student_Email box 
	student_email.focus_set() 

# Function for clearing the 
# contents of text entry boxes 
def clear(): 
	
	# clear the content of text entry box 
	surname.delete(0, END) 
	other_names.delete(0, END) 
	mat_reg_no.delete(0, END) 
	year_admitted.delete(0, END) 
	level.delete(0, END) 
	mobile_number.delete(0, END) 
	sex.delete(0, END) 
	address.delete(0, END) 
	state.delete(0, END) 
	lga.delete(0, END) 
	parents_guardian.delete(0, END) 
	parent_phone.delete(0, END)
	student_email.delete(0, END)


# Function to take data from GUI 
# window and write to an excel file 
def insert(): 
	
	# if user not fill any entry 
	# then print "empty input" 
	if (surname.get() == "" and
		other_names.get() == "" and
		mat_reg_no.get() == "" and
		year_admitted.get() == "" and
		level.get() == "" and
		mobile_number.get() == "" and
		sex.get() == "" and
		address.get() == "" and
		state.get() == "" and
		lga.get() == "" and
		parents_guardian.get() == "" and
		parent_phone.get() == "" and
		student_email.get() == ""): 
			
		print("empty input") 

	else: 

		# assigning the max row and max column 
		# value upto which data is written 
		# in an excel sheet to the variable 
		current_row = sheet.max_row 
		current_column = sheet.max_column 

		# get method returns current text 
		# as string which we write into 
		# excel spreadsheet at particular location 
		sheet.cell(row=current_row + 1, column=1).value = surname.get() 
		sheet.cell(row=current_row + 1, column=2).value = other_names.get() 
		sheet.cell(row=current_row + 1, column=3).value = mat_reg_no.get() 
		sheet.cell(row=current_row + 1, column=4).value = year_admitted.get() 
		sheet.cell(row=current_row + 1, column=5).value = level.get() 
		sheet.cell(row=current_row + 1, column=6).value = mobile_number.get() 
		sheet.cell(row=current_row + 1, column=7).value = sex.get()
		sheet.cell(row=current_row + 1, column=8).value = address.get() 
		sheet.cell(row=current_row + 1, column=9).value = state.get() 
		sheet.cell(row=current_row + 1, column=10).value = lga.get() 
		sheet.cell(row=current_row + 1, column=11).value = parents_guardian.get() 
		sheet.cell(row=current_row + 1, column=12).value = parent_phone.get() 
		sheet.cell(row=current_row + 1, column=13).value = student_email.get() 

		# save the file 
		wb.save('C:\\Users\\Mark Ekperi\\MyProjects\\RegistrationForm\\StudentRegForm.xlsx') 

		# set focus on the Surname box 
		surname.focus_set() 

		# call the clear() function 
		clear() 


# Driver code 
if __name__ == "__main__": 
	
	# create a GUI window 
	root = Tk() 

	# set the background colour of GUI window 
	root.configure(background='light blue') 

	# set the title of GUI window 
	root.title("Student's Registration form") 

	# set the configuration of GUI window 
	root.geometry("500x300") 

	excel() 

	# create a Form label 
	heading = Label(root, text="Department of Philosophy, Faculty of Arts, Niger Delta University", bg="light blue")

	# create a Surname label 
	surname = Label(root, text="Surname", bg="light blue") 

	# create a Other Names label 
	other_names = Label(root, text="Other Names", bg="light blue") 

	# create a Mat No label 
	mat_reg_no = Label(root, text="Matric No.", bg="light blue") 

	# create a Year Admitted lable 
	year_admitted = Label(root, text="Year Admitted", bg="light blue") 

	# create a Level label 
	level = Label(root, text="Level", bg="light blue") 

	# create a Mobile Number label 
	mobile_number = Label(root, text="Mobile Number", bg="light blue") 

	# create a Sex label 
	sex = Label(root, text="Sex", bg="light blue") 

	# create a address label 
	address = Label(root, text="Address", bg="light blue")

	# create a State label 
	state = Label(root, text="State", bg="light blue") 

	# create a LGA label 
	lga = Label(root, text="LGA", bg="light blue") 

	# create a Parents/Guardian lable 
	parents_guardian = Label(root, text="Parents", bg="light blue") 

	# create a Parent's Number label 
	parent_phone = Label(root, text="Parent's Phone", bg="light blue") 

	# create a Student Email label 
	student_email = Label(root, text="Student's Email", bg="light blue") 



	# grid method is used for placing 
	# the widgets at respective positions 
	# in table like structure . 
	heading.grid(row=0, column=1) 
	surname.grid(row=1, column=0) 
	other_names.grid(row=2, column=0) 
	mat_reg_no.grid(row=3, column=0) 
	year_admitted.grid(row=4, column=0) 
	level.grid(row=5, column=0) 
	mobile_number.grid(row=6, column=0) 
	sex.grid(row=7, column=0) 
	address.grid(row=8, column=0) 
	state.grid(row=9, column=0) 
	lga.grid(row=10, column=0) 
	parents_guardian.grid(row=11, column=0) 
	parent_phone.grid(row=12, column=0) 
	student_email.grid(row=13, column=0) 


	# create a text entry box 
	# for typing the information 
	surname = Entry(root) 
	other_names = Entry(root) 
	mat_reg_no = Entry(root) 
	year_admitted = Entry(root) 
	level = Entry(root) 
	mobile_number = Entry(root) 
	sex = Entry(root) 
	address = Entry(root) 
	state = Entry(root) 
	lga = Entry(root) 
	parents_guardian = Entry(root) 
	parent_phone = Entry(root) 
	student_email = Entry(root)

	# bind method of widget is used for 
	# the binding the function with the events 

	# whenever the enter key is pressed 
	# then call the focus1 function 
	surname.bind("<Return>", focus1) 

	# whenever the enter key is pressed 
	# then call the focus2 function 
	other_names.bind("<Return>", focus2) 

	# whenever the enter key is pressed 
	# then call the focus3 function 
	mat_reg_no.bind("<Return>", focus3) 

	# whenever the enter key is pressed 
	# then call the focus4 function 
	year_admitted.bind("<Return>", focus4) 

	# whenever the enter key is pressed 
	# then call the focus5 function 
	level.bind("<Return>", focus5) 

	# whenever the enter key is pressed 
	# then call the focus6 function 
	mobile_number.bind("<Return>", focus6) 

	sex.bind("<Return>", focus7)

	address.bind("<Return>", focus8)

	state.bind("<Return>", focus9)

	lga.bind("<Return>", focus10)

	parents_guardian.bind("<Return>", focus11)

	parent_phone.bind("<Return>", focus12)

	student_email.bind("<Return>", focus13)

	# grid method is used for placing 
	# the widgets at respective positions 
	# in table like structure . 
	surname.grid(row=1, column=1, ipadx=50) 
	other_names.grid(row=2, column=1, ipadx=50) 
	mat_reg_no.grid(row=3, column=1, ipadx=50) 
	year_admitted.grid(row=4, column=1, ipadx=50) 
	level.grid(row=5, column=1, ipadx=50) 
	mobile_number.grid(row=6, column=1, ipadx=50) 
	sex.grid(row=7, column=1, ipadx=50) 
	address.grid(row=8, column=1, ipadx=50) 
	state.grid(row=9, column=1, ipadx=50) 
	lga.grid(row=10, column=1, ipadx=50) 
	parents_guardian.grid(row=11, column=1, ipadx=50) 
	parent_phone.grid(row=12, column=1, ipadx=50) 
	student_email.grid(row=13, column=1, ipadx=50)

	# call excel function 
	excel() 

	# create a Submit Button and place into the root window 
	submit = Button(root, text="Submit", fg="Black", 
							bg="green", command=insert) 
	submit.grid(row=14, column=1) 

	# start the GUI 
	root.mainloop() 
